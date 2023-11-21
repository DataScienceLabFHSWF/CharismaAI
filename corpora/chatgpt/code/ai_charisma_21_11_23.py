import pandas as pd
import requests
from openpyxl import Workbook
import re

# Define constants
API_KEY = ""
INPUT_EXCEL_FILE = "steffi_yasser_fig_speech_v2.xlsx"
OUTPUT_EXCEL_FILE = "revised_4.xlsx"
N = 90



# Function to send a prompt to the ChatGPT API
def send_prompt_to_gpt(prompt):
    api_endpoint = "https://api.openai.com/v1/chat/completions"

    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json"
    }

    data = {
        "model": "gpt-3.5-turbo",
        
        "messages": [
            {"role": "system", "content": "You are a helpful assistant."},
            {"role": "user", "content": prompt}
        ]
    }

    response = requests.post(api_endpoint, json=data, headers=headers)
    return response.json()["choices"][0]["message"]["content"]

# Function to split response based on numbers
def split_response(response):
    response_lines = response.split('\n')
    response_dict = {}
    current_number = None

    for line in response_lines:
        match = re.match(r'^(\d+)\. (.+)', line)
        if match:
            current_number = match.group(1)
            response_dict[current_number] = [match.group(2)]
        elif current_number is not None:
            response_dict[current_number].append(line)

    return response_dict

# Load the Excel file excluding the first three rows
df = pd.read_excel(INPUT_EXCEL_FILE, skiprows=3)

# Initialize a dictionary to store prompts and responses for each command
command_data = {}

# Iterate through the first 'N' rows
for index, row in df.iterrows():
    if index >= N:
        break

    original_values = (row[0], row[1],row[2])
    prompt = f"{original_values[0]} and {original_values[1]}"
    
    # Print the current prompt being processed
    print(f"Processing prompt: {prompt}")
    
    response = send_prompt_to_gpt(prompt)
    command_data[original_values] = response

# Initialize a new workbook
wb = Workbook()

for i, (original_values, response) in enumerate(command_data.items(), start=1):
    response_dict = split_response(response)
    sheet_name = f"Command_{i}"

    # Create a new sheet
    ws = wb.create_sheet(title=sheet_name)

    # Set up the cells
    ws['A2'] = "Prompt"
    ws['A1'] = "Tactic"
    ws['B2'] = original_values[0]
    ws['B1'] = original_values[1]
    ws['A6'] = 'Comment'
    ws['A5'] = 'No'
    ws['B5'] = 'Results'
    ws['C5'] = 'Coding'
    ws['D5'] = 'Text'
    ws['D1'] = 'UID'
    ws['D2'] = original_values[2]
    # Populate the sheet with split responses and numbers to the left
    for identifier, responses in response_dict.items():
        for r in responses:
            ws.append([identifier, r])

    # Print a completion message
    print(f"Command {i} completed and saved in sheet: {sheet_name}")

# Remove the default sheet created and save the workbook
wb.remove(wb.active)
wb.save(OUTPUT_EXCEL_FILE)

print(f"Data has been processed and saved to {OUTPUT_EXCEL_FILE}")
